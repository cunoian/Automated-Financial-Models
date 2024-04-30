from openpyxl import load_workbook

#load data from 'Inputs.xslx'
source_workbook = load_workbook('Inputs.xlsx')
source_sheet = source_workbook['Inputs']

#select specific data from inputs and store in variables
company_title = source_sheet['C2'].value
share_price = source_sheet['F2'].value


yearminusfour = source_sheet['B6'].value
yearminusthree = source_sheet['B7'].value
yearminustwo = source_sheet['B8'].value
yearminusone = source_sheet['B9'].value
year = source_sheet['B10'].value




#open master excel file
master_workbook = load_workbook('DCF.xlsx')
master_IS = master_workbook['Income Statement']
master_NWC = master_workbook['NWC Calculation']
master_BS = master_workbook['Partial BS']
master_WACC = master_workbook['WACC']
master_FCF = master_workbook['DCFCF']
master_FCFTM = master_workbook['DCFCF Terminal Multiple']

#place inputs into appropriate cells in model

##input into income statement
master_IS['A1'].value = company_title
master_IS['C4'].value = yearminusfour
master_IS['D4'].value = yearminusthree
master_IS['E4'].value = yearminustwo
master_IS['F4'].value = yearminusone
master_IS['G4'].value = year

"Sales"
master_IS['C6'].value = source_sheet['C6'].value
master_IS['D6'].value = source_sheet['C7'].value
master_IS['E6'].value = source_sheet['C8'].value
master_IS['F6'].value = source_sheet['C9'].value
master_IS['G6'].value = source_sheet['C10'].value

"COGS"
master_IS['C8'].value = source_sheet['E6'].value
master_IS['D8'].value = source_sheet['E7'].value
master_IS['E8'].value = source_sheet['E8'].value
master_IS['F8'].value = source_sheet['E9'].value
master_IS['G8'].value = source_sheet['E10'].value

"SG&A"
master_IS['C11'].value = source_sheet['G6'].value
master_IS['D11'].value = source_sheet['G7'].value
master_IS['E11'].value = source_sheet['G8'].value
master_IS['F11'].value = source_sheet['G9'].value
master_IS['G11'].value = source_sheet['G10'].value

"R&D"
master_IS['C12'].value = source_sheet['I6'].value
master_IS['D12'].value = source_sheet['I7'].value
master_IS['E12'].value = source_sheet['I8'].value
master_IS['F12'].value = source_sheet['I9'].value
master_IS['G12'].value = source_sheet['I10'].value

"D&A"
master_IS['C14'].value = source_sheet['N6'].value
master_IS['D14'].value = source_sheet['N7'].value
master_IS['E14'].value = source_sheet['N8'].value
master_IS['F14'].value = source_sheet['N9'].value
master_IS['G14'].value = source_sheet['N10'].value

"Assumptions"
master_IS['H21'].value = source_sheet['D13'].value
master_IS['H22'].value = source_sheet['F13'].value
master_IS['H24'].value = source_sheet['H13'].value
master_IS['H25'].value = source_sheet['J13'].value
master_IS['H26'].value = source_sheet['L13'].value
master_IS['H29'].value = source_sheet['O13'].value

##input into NWC
"Accounts Receivable"
master_NWC['C6'].value = source_sheet['C17'].value
master_NWC['D6'].value = source_sheet['C18'].value
master_NWC['E6'].value = source_sheet['C19'].value
master_NWC['F6'].value = source_sheet['C20'].value
master_NWC['G6'].value = source_sheet['C21'].value

"Inventory"
master_NWC['C7'].value = source_sheet['D17'].value
master_NWC['D7'].value = source_sheet['D18'].value
master_NWC['E7'].value = source_sheet['D19'].value
master_NWC['F7'].value = source_sheet['D20'].value
master_NWC['G7'].value = source_sheet['D21'].value

"Other Current Assets"
master_NWC['C8'].value = source_sheet['E17'].value
master_NWC['D8'].value = source_sheet['E18'].value
master_NWC['E8'].value = source_sheet['E19'].value
master_NWC['F8'].value = source_sheet['E20'].value
master_NWC['G8'].value = source_sheet['E21'].value

"Accounts Payable"
master_NWC['C11'].value = source_sheet['F17'].value
master_NWC['D11'].value = source_sheet['F18'].value
master_NWC['E11'].value = source_sheet['F19'].value
master_NWC['F11'].value = source_sheet['F20'].value
master_NWC['G11'].value = source_sheet['F21'].value

"Other Non-Debt Liabilities"
master_NWC['C12'].value = source_sheet['G17'].value
master_NWC['D12'].value = source_sheet['G18'].value
master_NWC['E12'].value = source_sheet['G19'].value
master_NWC['F12'].value = source_sheet['G20'].value
master_NWC['G12'].value = source_sheet['G21'].value

"CAPX"
master_NWC['C28'].value = source_sheet['P17'].value
master_NWC['D28'].value = source_sheet['P18'].value
master_NWC['E28'].value = source_sheet['P19'].value
master_NWC['F28'].value = source_sheet['P20'].value
master_NWC['G28'].value = source_sheet['P21'].value
master_NWC['H30'].value = source_sheet['Q23'].value

##input into balance sheet
"Cash C5"
master_BS['C5'].value = source_sheet['H21'].value

"Accounts Receivable C6"
master_BS['C6'].value = source_sheet['C21'].value

"Inventory C7"
master_BS['C7'].value = source_sheet['D21'].value

"Other Current Assets C8"
master_BS['C8'].value = source_sheet['E21'].value

"PP&E C11"
master_BS['C11'].value = source_sheet['I21'].value

"Goodwill C12"
master_BS['C12'].value = source_sheet['R21'].value

"Other Long Term Assets C13"
master_BS['C13'].value = source_sheet['J21'].value

"Accounts Payable C17"
master_BS['C17'].value = source_sheet['F21'].value

"Other Current Liabilities C18"
master_BS['C18'].value = source_sheet['G21'].value

"Current Portion of Debt C19"
master_BS['C19'].value = source_sheet['K21'].value

"Long Term Debt C22"
master_BS['C22'].value = source_sheet['L21'].value

"Pension C23"
master_BS['C23'].value = source_sheet['M21'].value

"Other Long Term Liabilities C24"
master_BS['C24'].value = source_sheet['N21'].value

"Equity C27"
master_BS['C27'].value = source_sheet['O21'].value

"Shares Outstanding C32"
master_BS['C32'].value = source_sheet['K27'].value

##input into WACC 
"Beta B8"
master_WACC['B8'].value = source_sheet['E26'].value


"Risk Free Rate B6"
master_WACC['B6'].value = source_sheet['C26'].value

"Equity Risk Premium B7"
master_WACC['B7'].value = source_sheet['C27'].value

"Size Risk B10"
master_WACC['B10'].value = source_sheet['E27'].value


"Specific Risk B11"
master_WACC['B11'].value = source_sheet['G26'].value


"Cost of Debt D18"
master_WACC['D18'].value = source_sheet['G27'].value

"Market Cap B23"
master_WACC['B23'].value = source_sheet['I26'].value


"Total Debt B24"
master_WACC['B24'].value = source_sheet['I27'].value


##input into FCF
"Long Term Growth Assumption H6"
master_FCF['H6'].value = source_sheet['I2'].value

"Current Share Price C38"
master_FCF['C38'].value = source_sheet['F2'].value

##input into FCFTM
"Terminal Multiple"
master_FCFTM['H18'].value = source_sheet['K26'].value

"Long Term Growth Assumption H6"
master_FCFTM['H6'].value = source_sheet['I2'].value

"Current Share Price"
master_FCFTM['C35'].value = source_sheet['F2'].value

#save workbook
master_workbook.save('DCF.xlsx')
print("The model has been updated.")